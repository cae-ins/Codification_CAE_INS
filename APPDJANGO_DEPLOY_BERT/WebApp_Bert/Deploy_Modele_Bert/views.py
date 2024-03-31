import math
from django.shortcuts import render

from django.http import HttpResponse
import pandas as pd
from io import StringIO
from django.conf import settings
import os
from .mod_predict import predict_sic_code
from .control_file import VerificateurTexte
from .file_checker import verifier_fichier_csv
from transformers import DistilBertForSequenceClassification, DistilBertTokenizerFast
import pickle
import zipfile
import io
import tempfile
from django.http import JsonResponse
from django.middleware.csrf import get_token
from django.views.decorators.csrf import csrf_exempt
import json

from rest_framework.views import APIView
from rest_framework import status

from django.http import StreamingHttpResponse

#Les modules de classeur unique
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def start_codification(request, temp_dir):
   """
    Fonction permettant de faire la codification des fichiers CSV uploadés et de retourner la progression de la transformation.

    Args:
        request: Requête HTTP.
        temp_dir (str): Chemin vers le dossier temporaire contenant les fichiers uploadés.

    Returns:
        StreamingHttpResponse: Réponse HTTP en continu avec la progression de la transformation.

    Raises:
        Exception: Toutes les exceptions sont capturées et renvoyées dans la réponse.

    Note:
        Cette fonction utilise des techniques de streaming pour envoyer progressivement la progression
        du traitement au client.

    """
   def predict(temp_dir):

      """
         Fonction permettant de prédire la codification CITP des emplois à partir d'un modèle.

         Args:
            temp_dir (str): Chemin vers le dossier temporaire contenant les fichiers uploadés.

         Yields:
            str: Messages SSE (Server-Sent Events) indiquant l'état et la progression du traitement.

         Raises:
            Exception: Toutes les exceptions sont capturées et renvoyées dans la réponse.

         Note:
            Cette fonction utilise des techniques de streaming pour envoyer progressivement la progression
            du traitement au client.

      """
      #yield '{"statut": "error", "message": "Une erreur est apparue lors de l\'initialisation.", "progress": 0}'
      progress_data = {
            "statut": "in progress",
            "message": "Traitement en cours",
            "progress": 0
      }
      # Convertir le dictionnaire en chaîne JSON
      progress_json = json.dumps(progress_data)
      # Envoyer la chaîne JSON en tant que message SSE
      yield f"data: {progress_json}\n\n"

      try:

         # Recuperation du fichier comportant les details sur les fichiers a codifier
         with open(os.path.join(tempfile.gettempdir(), temp_dir, 'data.json'), 'r') as fichier_json:
            data = json.load(fichier_json)
         
         verificateur = VerificateurTexte()
         model_load_path = os.path.join(settings.STATICFILES_DIRS[0], 'Deploy_Modele_Bert', 'fine_tuned_model_runpod_distillbert')
         tokenizer_load_path = os.path.join(settings.STATICFILES_DIRS[0], 'Deploy_Modele_Bert', 'fine_tuned_tokenizer_runpod_distillbert')
         label_encoder_load_path =  os.path.join(settings.STATICFILES_DIRS[0], 'Deploy_Modele_Bert', 'label_encoder_runpod_distill_bert.pkl')

         with open(label_encoder_load_path, "rb") as f:
            label_encoder = pickle.load(f)
            f.close()

         model = DistilBertForSequenceClassification.from_pretrained(model_load_path)
         tokenizer = DistilBertTokenizerFast.from_pretrained(tokenizer_load_path)
         
         index_item = 1
         for item in data:
            # Recuperation des infos du fichier
            print(index_item)
            file_path = item["file_path"]
            file_name = item["name"].split('.')[0]
            colonne = item["colonne"]
            niveau = item["niveau"]
            '''
            # Vérifier si le fichier est un fichier CSV
            if file_path.lower().endswith('.csv') and os.path.isfile(file_path):
                
               with open(file_path, 'r') as file:
                  
                  # Déplacer le curseur de lecture au début du fichier
                  file.seek(0)
                  f_input_excel1 = pd.read_csv(file, sep=";", encoding="latin-1")
                  # Chemin de sortie du fichier Excel
                  xlsx_file_path = os.path.splitext(file_path)[0] + '.xlsx'
                  # Écrire les données dans un fichier Excel
                  f_input_excel1.to_excel(xlsx_file_path, index=False)
                  # Supprimer le fichier CSV d'origine
                  os.remove(file_path)
             '''
            #Lecture du fichier a codifier
            with open(file_path, 'r', encoding='latin-1') as input_file:
                
               print(input_file)
               #Verification du fichier à codifier
               control_file = verifier_fichier_csv(input_file)
               if control_file['status'] == 'error':
                  yield f"data: {json.dumps({'statut': 'error', 'message': control_file['message'], 'progress': index_item * 100 / len(data)})}\n\n"

               else:

                  input_file.seek(0)
                  #f_input = pd.read_csv(input_file, sep=";", encoding="latin-1")
                  f_input = pd.read_excel(file_path, engine='openpyxl')

                  caracteres_errones = [] 
                  for index, row in f_input.iterrows():
                     text = str(row['libelle'])
                     if verificateur.verifie_longueur(text) or verificateur.verifie_caractere_unique(text) or \
                     verificateur.verifie_trois_successifs(text) or verificateur.verifie_chiffres_uniquement(text):
                        caracteres_errones.append(text)
                        f_input = f_input.drop(index)
                     else:
                        predictions = predict_sic_code(text, model, tokenizer, label_encoder)
                        clef = []
                        valeur = []
                        dict_pred = {}
                        for i, (sic_code, certainty) in enumerate(predictions, 1):
                              clef.append(sic_code)
                              valeur.append(certainty)
                        for cle, valeur in zip(clef, valeur):
                              dict_pred[cle] = valeur
                        cle_max = max(dict_pred, key=dict_pred.get)
                        f_input.loc[index, 'Code'] = cle_max
                        f_input.loc[index, 'Vraisemblance'] = dict_pred[cle_max]

                     progress = math.floor(((index_item - 1) / len(data)) * 100 + (100 / len(data)) * index / len(f_input))
                     yield f"data: {json.dumps({'statut': 'in progress', 'message': 'Traitement en cours', 'progress': progress})}\n\n"
                  '''
                  #Creation du dossier des outputs de la codification
                  output_dir = os.path.join(tempfile.gettempdir(), temp_dir, 'output')
                  if not os.path.exists(output_dir):
                     os.makedirs(output_dir)
                  print(output_dir)
                  df_errone = pd.DataFrame({"libelle_errone": caracteres_errones})
                  errone_file_path = os.path.join(output_dir, f'errone_data_{file_name}.xlsx')
                  #df_errone.to_csv(errone_file_path, sep=';', index=False)
                  df_errone.to_excel(errone_file_path, index=False)
                  transformed_file_path = os.path.join(output_dir, f'transformed_data_{file_name}.xlsx')
                  #f_input.to_csv(transformed_file_path, sep=';', index=False)
                  f_input.to_excel(transformed_file_path, index=False)
                  '''
                  #Creation du dossier des outputs de la codification
                  output_dir = os.path.join(tempfile.gettempdir(), temp_dir, 'output')
                  if not os.path.exists(output_dir):
                     os.makedirs(output_dir)
                  
                  print(output_dir)

                  df_errone = pd.DataFrame({"libelle_errone": caracteres_errones})
                  # Créer un classeur Excel
                  wb = Workbook()

                  # Créer une feuille pour les données erronées
                  ws_errone = wb.active
                  ws_errone.title = 'erronee_data'
                  for r in dataframe_to_rows(df_errone, index=False, header=True):
                     ws_errone.append(r)

                  # Créer une feuille pour les données transformées
                  ws_transformed = wb.create_sheet(title='transformed_data')
                  for r in dataframe_to_rows(f_input, index=False, header=True):
                     ws_transformed.append(r)
                  
                  # Vérifier si la feuille par défaut 'Sheet' existe
                  if 'Sheet' in wb.sheetnames:
                     # Supprimer la feuille par défaut (Sheet) si elle existe
                     default_sheet = wb['Sheet']
                     wb.remove(default_sheet)
         
                  # Enregistrer le classeur dans un fichier
                  combined_file_path = os.path.join(output_dir, f'Data_Codif_{file_name}.xlsx')
                  wb.save(combined_file_path)
                  


               input_file.close()
            index_item += 1

      except Exception as e:
         yield f"data: {json.dumps({'statut': 'error', 'message': f'Une erreur est survenue: {str(e)}', 'progress': 0})}\n\n"

      yield f"data: {json.dumps({'statut': 'Complete', 'message': 'Traitement terminé avec success', 'progress': 100})}\n\n"

   response = StreamingHttpResponse(predict(temp_dir), content_type="text/event-stream")
   response['Cache-Control'] = 'no-cache'
   return response


def get_csrf_token(request):
    token = get_token(request)
    return JsonResponse({'csrfToken': token})

class UploadFiles(APIView):
   """
    Classe API permettant de gérer le téléchargement de fichiers.

    """
   def post(self, request, format=None):
        """
        Méthode POST pour télécharger les fichiers et les détails associés.

        Args:
            request (Request): Requête HTTP contenant les fichiers à télécharger.
            format (str): Format de la réponse HTTP. Par défaut, None.

        Returns:
            JsonResponse: Réponse JSON contenant le statut de la requête et des détails sur les fichiers téléchargés.

        """
        data_details = request.data.get('data_details')
        data = []
        
        # Créer un répertoire temporaire
        temp_dir = tempfile.mkdtemp(prefix="codif_")
        if data_details:
            #recuperation des details inclus dans la requete post
            details = json.loads(data_details)
            for detail in details:
                file_index = detail.get("index")
                uploaded_file = request.FILES.get(file_index)
                if uploaded_file:
                    #detail["file"] = uploaded_file
                    file_name = uploaded_file.name
                    file_path = os.path.join(temp_dir, file_name)
                    detail["file_path"] = file_path
                    data.append(detail)

                    #Enregistrement dans le fichier temporaire des fichiers uploadés.
                    with open(file_path, 'wb') as destination:
                       for chunk in uploaded_file.chunks():
                           destination.write(chunk)
                else:
                    return JsonResponse({
                        "statut": 'error',
                        "message": "Fichier manquant pour l'index {}".format(detail.get("name")),
                    }, status=status.HTTP_400_BAD_REQUEST)
        
        detail_file_path = os.path.join(temp_dir, f"data.json")
        with open(detail_file_path, 'w') as detail_file:
            json.dump(data, detail_file)


        return JsonResponse({
            "temp_dir": temp_dir.split("\\")[-1],  # Nom du dossier temporaire
            "statut": 'succes',
            "message": "Fichiers uploader avec succès"
        }, status=status.HTTP_201_CREATED)

#Fonction pour la transformation du fichier
@csrf_exempt
def predict(request):
 
      try:  
            #L'objet contenant les critères de verification
            verificateur = VerificateurTexte()

            # Créer des répertoires temporaires
            temp_dir = tempfile.mkdtemp(prefix="codif_")

            # Stocker les valeurs dans la session
            request.session['temp_dir'] = temp_dir
            
            # Emplacements où le modèle et le tokenizer ont été sauvegardés
            model_load_path = os.path.join(settings.STATICFILES_DIRS[0], 'Deploy_Modele_Bert', 'fine_tuned_model_runpod_distillbert')
            tokenizer_load_path = os.path.join(settings.STATICFILES_DIRS[0], 'Deploy_Modele_Bert', 'fine_tuned_tokenizer_runpod_distillbert')
            label_encoder_load_path =  os.path.join(settings.STATICFILES_DIRS[0], 'Deploy_Modele_Bert', 'label_encoder_runpod_distill_bert.pkl')

            with open(label_encoder_load_path, "rb") as f:
               label_encoder = pickle.load(f)

            # Charger le modèle DistilBERT
            model = DistilBertForSequenceClassification.from_pretrained(model_load_path)
            # Charger le tokenizer associé
            tokenizer = DistilBertTokenizerFast.from_pretrained(tokenizer_load_path)

            if request.method == 'POST' and request.FILES :
               
               input_file = request.FILES.get("file")

               control_file = verifier_fichier_csv(input_file)

               if(control_file['status'] == 'error'):
                    
                  return JsonResponse({
                  #"temp_dir" : temp_dir.split("\\")[-1],
                  "statut" : 'error',
                  "message" : control_file['message']
               })

               else:
                  input_file.seek(0)

                  file_content = input_file.read().decode('latin-1')
                  
                  # Utiliser StringIO pour créer un objet fichier virtuel
                  file_object = StringIO(file_content)
                  
                  # Lire le contenu CSV dans un DataFrame
                  f_input = pd.read_csv(file_object, sep=";", encoding="latin-1")

                  j=-1
                  caracteres_errones = []
                  for index, row in f_input.iterrows():
                     
                     j+=1
                     text = str(row['libelle'])
                     #Condition de verification des caractères du libellé
                     if verificateur.verifie_longueur(text) or verificateur.verifie_caractere_unique(text) or \
                     verificateur.verifie_trois_successifs(text) or verificateur.verifie_chiffres_uniquement(text):
                        
                        caracteres_errones.append(text)
                        f_input = f_input.drop(index)

                     else:
               
                        # Prédiction avec le modèle et le tokenizer chargés
                        predictions = predict_sic_code(text, model, tokenizer, label_encoder)
                        clef =[]
                        valeur =[]
                        dict_pred = {}
                        for i, (sic_code, certainty) in enumerate(predictions, 1):
                           clef.append(sic_code)
                           valeur.append(certainty)
                        for cle, valeur in zip(clef, valeur):
                           dict_pred[cle] = valeur
                        cle_max = max(dict_pred, key=dict_pred.get)
                        f_input.loc[j, 'Code'] = cle_max
                        f_input.loc[j, 'Vraisemblance'] = dict_pred[cle_max]

                  #Exportation du fichier contenant des données érronées sur le serveur
                  df_errone = pd.DataFrame({"libelle_errone": caracteres_errones})
                  #errone_file_path = os.path.join(settings.MEDIA_ROOT, 'transformed_files', 'errone_data.csv')
                  #df_errone.to_csv(errone_file_path, sep =';', index=False)
                  errone_file_path = os.path.join(temp_dir, 'errone_data.xlsx')
                  df_errone.to_csv(errone_file_path, sep=';', index=False)
                  
                  #Exportation du fichier contenant des données transformées sur le serveur
                  #transformed_file_path = os.path.join(settings.MEDIA_ROOT, 'transformed_files', 'transformed_data.csv')
                  #f_input.to_csv(transformed_file_path, sep =';', index=False)
                  
                  transformed_file_path = os.path.join(temp_dir, 'transformed_data.xlsx')
                  f_input.to_csv(transformed_file_path, sep=';', index=False)

                  #Renvoyer la page pour telecharger le fichier
                  return JsonResponse({
                     "temp_dir" : temp_dir.split("\\")[-1],
                     "statut" : 'succes',
                     "message" : "Traitement effectuer avec succes"
                  })
            

      except Exception as e:
          
            print("Erreur dans le fichier:", e)
            
   
#Fonction qui retourne la page d'accueil
def index(request):

   #Renvoyer la page pour charger le fichier
   #img_path = os.path.join(settings.MEDIA_ROOT,'images','logo-ins.png')
   #img_path = os.path.join(settings.STATICFILES_DIRS[0], 'Deploy_Modele_Bert', 'images','logo-ins.png')
   return render(request,'index.html')

def get_models_details(request):
   with open(os.path.join(settings.STATICFILES_DIRS[0], 'Deploy_Modele_Bert', 'config.json'), 'r') as fichier_json:
      data = json.load(fichier_json)
   
   print(data)

   return JsonResponse({
      "data" : data
   })

#Fonction qui retourne la page de téléchargement
def download_page(request):
    return render(request, 'Deploy_Modele_Bert/download_page.html')

#Fonction pour le telechargement du fichier transformé
def download_transformed_csv(request, temp_dir):
    
    print(os.path.join(tempfile.gettempdir(), temp_dir))
    temp_dir = os.path.join(tempfile.gettempdir(), temp_dir, 'output')
    print(temp_dir)
    # Récupérer les valeurs depuis la session
    # Chemins vers les fichiers transformés
    #transformed_file_path = os.path.join(settings.MEDIA_ROOT, 'transformed_files', 'transformed_data.csv')
    guide_file_path = os.path.join(settings.MEDIA_ROOT, 'guide', 'guide.xlsx')
    #transformed_file_path = os.path.join(temp_dir, 'transformed_data.csv')
    #transformed_file_path = os.path.join(temp_dir, 'transformed_data.csv')
    #errone_file_path = os.path.join(settings.MEDIA_ROOT, 'transformed_files', 'errone_data.csv')
    #errone_file_path = os.path.join(temp_dir, 'errone_data.csv')
    '''
    # Initialise un nouveau classeur Excel
    wb = Workbook()

    # Ajoutez chaque fichier Excel en tant que feuille
    for file_name in os.listdir(temp_dir):
        if file_name.endswith('.xlsx'):  # Vérifiez si le fichier est au format Excel
            sheet_name = os.path.splitext(file_name)[0]  # Utilisez le nom du fichier comme nom de feuille
            sheet = wb.create_sheet(title=sheet_name)  # Créez une nouvelle feuille
            df = pd.read_excel(os.path.join(temp_dir, file_name))  # Lisez le fichier Excel dans un DataFrame
            for r in dataframe_to_rows(df, index=False, header=True):  # Ajoutez les données DataFrame à la feuille
                sheet.append(r)

    # Ajoutez le fichier guide en tant que feuille
    guide_sheet = wb.create_sheet(title='Guide')
    guide_df = pd.read_excel(guide_file_path)
    for r in dataframe_to_rows(guide_df, index=False, header=True):
        guide_sheet.append(r)

    # Supprimez la première feuille par défaut (Sheet)
    default_sheet = wb['Sheet']
    wb.remove(default_sheet)

    # Enregistrez le classeur dans un buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)

    # Retournez le buffer Excel pour téléchargement
    excel_buffer.seek(0)
    response = HttpResponse(excel_buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=combined_excel.xlsx'
    return response

    '''
    # Récupérer la liste des fichiers dans le dossier temporaire
    file_names = os.listdir(temp_dir)

    # Créez un objet Zip
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
        # Ajoutez les fichiers au zip
         # Ajoutez tous les fichiers au zip
        for file_name in file_names:
            file_path = os.path.join(temp_dir, file_name)
            zip_file.write(file_path, arcname=file_name)
        zip_file.write(guide_file_path, arcname='guide.xlsx')


    #Suppression des repertoires temporaires
    #shutil.rmtree(temp_dir)

    # Répondez avec le contenu du zip
    response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename=transformed_files.zip'

    return response
    